from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from datetime import date
import math

YELLOW   = "FEDE07"
DARK     = "1A2332"
MID_GRAY = "F2F2F2"
LIGHT    = "FFFFFF"
BORDER_C = "CCCCCC"

def side(style="thin", color=BORDER_C):
    return Side(style=style, color=color)

def all_border(style="thin"):
    s = side(style)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, cell_ref, value=None, bold=False, size=10, color="000000",
               bg=None, align="right", wrap=False, border=True, num_fmt=None):
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", bold=bold, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = all_border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def build_annex(data: dict, output_path: str):
    tn    = data["tenant_name"]
    unit  = data["unit"]
    proj  = data["project"]
    ls    = data["lease_start"]
    ny    = data["num_years"]
    bm    = data["base_monthly"]
    esc   = data["escalation"]
    scm1  = data["sc_monthly_y1"]
    sc_esc= data.get("sc_escalation", esc)
    vr    = data.get("vat_rent", 0.01)
    vs    = data.get("vat_sc", 0.14)
    rs_yr = data.get("revenue_share_years", 0)
    mkt_rate = data.get("marketing_rate", 0.0)

    # Annual summary data (used for summary sheet, deposit, marketing)
    years = []
    for y in range(ny):
        yr_start = ls + relativedelta(years=y)
        yr_end   = ls + relativedelta(years=y+1) - relativedelta(days=1)
        is_rs    = (y < rs_yr)
        rent_monthly = 0.0 if is_rs else bm * ((1 + esc) ** y)
        sc_monthly   = scm1 * ((1 + sc_esc) ** max(0, y - rs_yr))
        mkt_annual   = round(rent_monthly * 12 * mkt_rate, 2)
        years.append({
            "y":           y + 1,
            "label":       f"Year {y+1}",
            "period":      f"{yr_start.strftime('%d/%m/%Y')} – {yr_end.strftime('%d/%m/%Y')}",
            "is_rs":       is_rs,
            "rent_m":      rent_monthly,
            "rent_ann":    round(rent_monthly * 12, 2),
            "sc_m":        sc_monthly,
            "sc_ann":      round(sc_monthly * 12, 2),
            "vat_rent_m":  round(rent_monthly * vr, 2),
            "vat_rent_ann":round(rent_monthly * vr * 12, 2),
            "vat_sc_m":    round(sc_monthly * vs, 2),
            "vat_sc_ann":  round(sc_monthly * vs * 12, 2),
            "mkt_ann":     mkt_annual,
            "vat_mkt_ann": round(mkt_annual * vs, 2),
        })

    # Monthly installment data
    monthly_rows = []
    for y_idx, y in enumerate(years):
        yr_start = ls + relativedelta(years=y_idx)
        for m in range(12):
            due_date = yr_start + relativedelta(months=m)
            monthly_rows.append({
                "installment": y_idx * 12 + m + 1,
                "due_date":    due_date,
                "year_label":  y["label"],
                "year_num":    y["y"],
                "is_rs":       y["is_rs"],
                "rent_m":      y["rent_m"],
                "vat_rent_m":  y["vat_rent_m"],
                "rent_total_m":round(y["rent_m"] + y["vat_rent_m"], 2),
                "sc_m":        y["sc_m"],
                "vat_sc_m":    y["vat_sc_m"],
                "sc_total_m":  round(y["sc_m"] + y["vat_sc_m"], 2),
            })

    first_rent_year = next((y for y in years if not y["is_rs"]), years[0])
    deposit = (first_rent_year["rent_m"] + first_rent_year["sc_m"]) * 3

    wb = Workbook()

    # ──────────────────────────────────────────────────────────────────
    # Sheet 1: Monthly Payment Schedule
    # Columns: # | Due Date | Year | Base Amount | VAT Rate % | VAT Amount | Total Incl VAT
    # ──────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Payment Schedule"
    ws.column_dimensions["A"].width = 6    # #
    ws.column_dimensions["B"].width = 14   # Due Date
    ws.column_dimensions["C"].width = 16   # Year
    ws.column_dimensions["D"].width = 18   # Base Amount
    ws.column_dimensions["E"].width = 12   # VAT Rate %
    ws.column_dimensions["F"].width = 16   # VAT Amount
    ws.column_dimensions["G"].width = 20   # Total Incl VAT

    NUM_COLS = 7

    def merge_header(row, title, bg=DARK, fg=LIGHT, size=14, height=32):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws["A" + str(row)]
        c.value = title
        c.font  = Font(name="Arial", bold=True, size=size, color=fg)
        c.fill  = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = height

    def section_header(row, title):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row, 1, title)
        c.font  = Font(name="Arial", bold=True, size=10, color=LIGHT)
        c.fill  = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20

    def col_header(row, cols):
        for j, txt in enumerate(cols, 1):
            c = ws.cell(row, j, txt)
            c.font  = Font(name="Arial", bold=True, size=9, color=DARK)
            c.fill  = PatternFill("solid", fgColor=MID_GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = all_border()
        ws.row_dimensions[row].height = 28

    def year_group_header(row, label, period):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row, 1, f"{label}   |   {period}")
        c.font  = Font(name="Arial", bold=True, size=9, color=DARK)
        c.fill  = PatternFill("solid", fgColor="D9E8F5")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = all_border()
        ws.row_dimensions[row].height = 18

    _row_alt = [0]
    def monthly_data_row(row, cols):
        _row_alt[0] += 1
        bg = LIGHT if _row_alt[0] % 2 == 1 else MID_GRAY
        for j, val in enumerate(cols, 1):
            c = ws.cell(row, j)
            c.value  = val
            c.font   = Font(name="Arial", size=9, color="000000")
            c.fill   = PatternFill("solid", fgColor=bg)
            c.border = all_border()
            if isinstance(val, (float, int)) and not isinstance(val, bool):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, date):
                c.number_format = 'DD/MM/YYYY'
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 17

    def year_subtotal_row(row, label, base_ann, vat_ann, total_ann):
        vals = [None, None, label, base_ann, None, vat_ann, total_ann]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row, j)
            c.value  = val
            c.font   = Font(name="Arial", bold=True, size=9, color=DARK)
            c.fill   = PatternFill("solid", fgColor="E8F4E8")
            c.border = all_border()
            if isinstance(val, (float, int)) and not isinstance(val, bool):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18

    def grand_total_row(row, label, base_total, vat_total, grand_total):
        vals = [None, None, label, base_total, None, vat_total, grand_total]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row, j)
            c.value  = val
            c.font   = Font(name="Arial", bold=True, size=9, color=DARK)
            c.fill   = PatternFill("solid", fgColor=YELLOW)
            c.border = all_border()
            if isinstance(val, (float, int)) and not isinstance(val, bool):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20

    # Main title
    merge_header(1, "Financial Annex — Monthly Payment Schedule", DARK, LIGHT, 14, 32)
    merge_header(2, f"{tn}  |  {unit}  |  {proj}", YELLOW, DARK, 11, 22)
    ws.merge_cells("A3:G3")
    ws.row_dimensions[3].height = 8

    r = 4

    # ── Section 1: Rental Value ──
    vr_pct = f"{vr*100:g}%"
    section_header(r, f"1. Rental Value Schedule  —  VAT Rate: {vr_pct}")
    r += 1
    col_header(r, ["#", "Due Date", "Year", "Base Rent (EGP)", f"VAT Rate",
                   "VAT Amount (EGP)", "Total Incl. VAT (EGP)"])
    r += 1

    _row_alt[0] = 0
    rent_grand_base = 0.0
    rent_grand_vat  = 0.0
    rent_grand_tot  = 0.0

    for y_idx, y in enumerate(years):
        yr_start = ls + relativedelta(years=y_idx)
        yr_end   = ls + relativedelta(years=y_idx+1) - relativedelta(days=1)
        year_group_header(r, y["label"],
                          f"{yr_start.strftime('%d/%m/%Y')} – {yr_end.strftime('%d/%m/%Y')}")
        r += 1
        yr_base = 0.0; yr_vat = 0.0; yr_tot = 0.0
        for m in range(12):
            due_date = yr_start + relativedelta(months=m)
            inst_num = y_idx * 12 + m + 1
            base_v   = y["rent_m"]
            vat_v    = y["vat_rent_m"]
            tot_v    = round(base_v + vat_v, 2)
            monthly_data_row(r, [inst_num, due_date, y["label"], base_v, vr_pct, vat_v, tot_v])
            yr_base += base_v; yr_vat += vat_v; yr_tot += tot_v
            r += 1
        year_subtotal_row(r, f"{y['label']} Subtotal",
                          round(yr_base, 2), round(yr_vat, 2), round(yr_tot, 2))
        rent_grand_base += yr_base; rent_grand_vat += yr_vat; rent_grand_tot += yr_tot
        r += 1

    grand_total_row(r, "GRAND TOTAL",
                    round(rent_grand_base, 2), round(rent_grand_vat, 2), round(rent_grand_tot, 2))
    r += 2

    # ── Section 2: Service Charges ──
    vs_pct = f"{vs*100:g}%"
    section_header(r, f"2. Service Charges Schedule  —  VAT Rate: {vs_pct}")
    r += 1
    col_header(r, ["#", "Due Date", "Year", "Base SC (EGP)", f"VAT Rate",
                   "VAT Amount (EGP)", "Total Incl. VAT (EGP)"])
    r += 1

    _row_alt[0] = 0
    sc_grand_base = 0.0
    sc_grand_vat  = 0.0
    sc_grand_tot  = 0.0

    for y_idx, y in enumerate(years):
        yr_start = ls + relativedelta(years=y_idx)
        yr_end   = ls + relativedelta(years=y_idx+1) - relativedelta(days=1)
        year_group_header(r, y["label"],
                          f"{yr_start.strftime('%d/%m/%Y')} – {yr_end.strftime('%d/%m/%Y')}")
        r += 1
        yr_base = 0.0; yr_vat = 0.0; yr_tot = 0.0
        for m in range(12):
            due_date = yr_start + relativedelta(months=m)
            inst_num = y_idx * 12 + m + 1
            base_v   = y["sc_m"]
            vat_v    = y["vat_sc_m"]
            tot_v    = round(base_v + vat_v, 2)
            monthly_data_row(r, [inst_num, due_date, y["label"], base_v, vs_pct, vat_v, tot_v])
            yr_base += base_v; yr_vat += vat_v; yr_tot += tot_v
            r += 1
        year_subtotal_row(r, f"{y['label']} Subtotal",
                          round(yr_base, 2), round(yr_vat, 2), round(yr_tot, 2))
        sc_grand_base += yr_base; sc_grand_vat += yr_vat; sc_grand_tot += yr_tot
        r += 1

    grand_total_row(r, "GRAND TOTAL",
                    round(sc_grand_base, 2), round(sc_grand_vat, 2), round(sc_grand_tot, 2))
    r += 2

    # ── Section 3: Marketing (if applicable) ──
    if mkt_rate > 0:
        section_header(r, "3. Marketing Charges")
        r += 1
        col_header(r, ["Year", "Period", "Base Annual Rent", "Marketing Rate",
                        "Annual Marketing Fee", "VAT (14%)", "Total (incl VAT)", "Payment Due"])
        # widen cols for 8-col section
        ws.column_dimensions["H"].width = 20
        r += 1
        _row_alt[0] = 0
        for y in years:
            total_ann = round(y["mkt_ann"] + y["vat_mkt_ann"], 2)
            yr_label  = f"First month of Year {y['y']}"
            for j, val in enumerate([y["label"], y["period"],
                                      round(y["rent_ann"],2), f"{mkt_rate*100:.0f}%",
                                      y["mkt_ann"], y["vat_mkt_ann"],
                                      total_ann, yr_label], 1):
                c = ws.cell(r, j)
                c.value  = val
                c.font   = Font(name="Arial", size=9)
                c.fill   = PatternFill("solid", fgColor=LIGHT)
                c.border = all_border()
                if isinstance(val, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 18
            r += 1
        tot_mkt_ann   = sum(y["mkt_ann"] for y in years)
        tot_vat_mkt   = sum(y["vat_mkt_ann"] for y in years)
        tot_mkt_total = tot_mkt_ann + tot_vat_mkt
        for j, val in enumerate(["Total", "", "", "", tot_mkt_ann, tot_vat_mkt, tot_mkt_total, ""], 1):
            c = ws.cell(r, j)
            c.value  = val
            c.font   = Font(name="Arial", bold=True, size=9, color=DARK)
            c.fill   = PatternFill("solid", fgColor=YELLOW)
            c.border = all_border()
            if isinstance(val, float):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 2
        deposit_label = "4. Security Deposit"
    else:
        deposit_label = "3. Security Deposit"

    # ── Security Deposit ──
    section_header(r, deposit_label)
    r += 1
    dep_rent = first_rent_year["rent_m"] * 3
    dep_sc   = first_rent_year["sc_m"]   * 3
    rows_dep = [
        ("3 Months Rent", dep_rent),
        ("3 Months Service Charge", dep_sc),
        ("Total Security Deposit", deposit),
    ]
    for label, val in rows_dep:
        is_total = label.startswith("Total")
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(r, 1, label)
        c.font   = Font(name="Arial", size=9, bold=is_total)
        c.fill   = PatternFill("solid", fgColor=(YELLOW if is_total else LIGHT))
        c.border = all_border()
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"F{r}:G{r}")
        c2 = ws.cell(r, 6, val)
        c2.font   = Font(name="Arial", size=9, bold=is_total)
        c2.fill   = PatternFill("solid", fgColor=(YELLOW if is_total else LIGHT))
        c2.border = all_border()
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    c = ws.cell(r, 1)
    c.value = (f"All amounts in Egyptian Pounds (EGP)  |  "
               f"Rent subject to annual escalation of {esc*100:g}% per annum")
    c.font  = Font(name="Arial", size=8, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ──────────────────────────────────────────────────────────────────
    # Sheet 2: Annual Summary
    # ──────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 25

    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value = "Financial Annex Summary"
    c.font  = Font(name="Arial", bold=True, size=13, color=LIGHT)
    c.fill  = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:C2")
    c = ws2["A2"]
    c.value = f"{tn}  |  {unit}  |  {proj}"
    c.font  = Font(name="Arial", bold=True, size=11, color=DARK)
    c.fill  = PatternFill("solid", fgColor=YELLOW)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 22

    info_rows = [
        ("Tenant", tn),
        ("Unit", unit),
        ("Project", proj),
        ("Lease Start", ls.strftime("%d %B %Y")),
        ("Lease Term", f"{ny} Years"),
        ("Escalation", f"{esc*100:g}% per annum"),
        ("VAT on Rent", f"{vr*100:.1f}%"),
        ("VAT on Service Charge", f"{vs*100:.0f}%"),
        ("Security Deposit", f"EGP {deposit:,.0f}"),
    ] + ([("Marketing Rate", f"{mkt_rate*100:.1f}% of annual rent")] if mkt_rate > 0 else [])

    r2 = 4
    for label, val in info_rows:
        c = ws2.cell(r2, 1, label)
        c.font   = Font(name="Arial", bold=True, size=9)
        c.fill   = PatternFill("solid", fgColor=MID_GRAY)
        c.border = all_border()
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.merge_cells(f"B{r2}:C{r2}")
        c2 = ws2.cell(r2, 2, val)
        c2.font   = Font(name="Arial", size=9)
        c2.border = all_border()
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r2].height = 18
        r2 += 1

    r2 += 1
    has_mkt = mkt_rate > 0
    headers = ["Year", "Period", "Monthly Rent", "Annual Rent", "Monthly SC", "Annual SC"] + \
              (["Monthly Marketing", "Annual Marketing"] if has_mkt else []) + \
              ["Total Monthly (incl VAT)", "Total Annual (incl VAT)"]
    for j, h in enumerate(headers, 1):
        ws2.column_dimensions[get_column_letter(j)].width = 18
        c = ws2.cell(r2, j, h)
        c.font  = Font(name="Arial", bold=True, size=9, color=DARK)
        c.fill  = PatternFill("solid", fgColor=MID_GRAY)
        c.border = all_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[r2].height = 28
    r2 += 1

    for y in years:
        label = f"Year {y['y']}" + (" (Rev. Share)" if y["is_rs"] else "")
        total_ann = round(y["rent_ann"] + y["sc_ann"] + y["vat_rent_ann"] + y["vat_sc_ann"] +
                          (y["mkt_ann"] + y["vat_mkt_ann"] if has_mkt else 0), 2)
        total_m   = round(total_ann / 12, 2)
        mkt_cols  = [y["mkt_ann"], y["vat_mkt_ann"], round(y["mkt_ann"]+y["vat_mkt_ann"],2)] if has_mkt else []
        row_data  = [label, y["period"], y["rent_m"], y["rent_ann"],
                     y["sc_m"], y["sc_ann"]] + mkt_cols + [total_m, total_ann]
        for j, val in enumerate(row_data, 1):
            c = ws2.cell(r2, j, val)
            c.font   = Font(name="Arial", size=9)
            c.border = all_border()
            c.alignment = Alignment(horizontal="right" if isinstance(val, float) else "center",
                                    vertical="center")
            if isinstance(val, float):
                c.number_format = '#,##0.00'
        ws2.row_dimensions[r2].height = 18
        r2 += 1

    tot_all_ann = sum(y["rent_ann"]+y["sc_ann"]+y["vat_rent_ann"]+y["vat_sc_ann"] for y in years)
    for j, val in enumerate(["TOTAL", "", "", "", "", "", "", tot_all_ann], 1):
        c = ws2.cell(r2, j, val)
        c.font  = Font(name="Arial", bold=True, size=9)
        c.fill  = PatternFill("solid", fgColor=YELLOW)
        c.border = all_border()
        c.alignment = Alignment(horizontal="right" if isinstance(val, float) else "center",
                                vertical="center")
        if isinstance(val, float):
            c.number_format = '#,##0.00'
    ws2.row_dimensions[r2].height = 20

    wb.save(output_path)
    return years, deposit

if __name__ == "__main__":
    test_data = {
        "tenant_name": "Cilantro",
        "unit": "F01",
        "project": "Giza Zoo Commercial Destination",
        "lease_start": date(2025, 11, 1),
        "num_years": 5,
        "base_monthly": 150_000,
        "escalation": 0.10,
        "sc_monthly_y1": 30_000,
        "sc_escalation": 0.10,
        "vat_rent": 0.01,
        "vat_sc": 0.14,
        "revenue_share_years": 0,
    }
    years, deposit = build_annex(test_data, "/tmp/test_annex.xlsx")
    print("\n── Rent Schedule ──")
    for y in years:
        print(f"  Y{y['y']} | Rent/mo: {y['rent_m']:>12,.0f}")
    print(f"  Security Deposit: EGP {deposit:,.0f}")
